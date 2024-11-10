import pandas as pd
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

current_dir = os.getcwd()
path = f"{current_dir}/Companies"
final_quaters = []
quater_Q = [""]
final_data_list = []
k = True
def prune_data(sector_name,company_name,file):
    global current_dir,quater_Q,final_quaters,k
    # print(current_dir)
    
    file_name = f"{current_dir}/Companies/{sector_name}/{company_name}/{file}"
    
    df = pd.read_excel(file_name)

    col_names = list(df.columns)
    #print(col_names)
    
    quaters = col_names[1:]
    # quaters.reverse()
    # print((quaters))
    j = 1
    
    if k :
        for i in range(0,(int(len(quaters)/5))):
            qua = quaters[(i*5) : (i*5) + 5]
            qua.reverse()
            # print(qua,"\n")
            for i in qua:
                quater_Q.append(f"Q{j}")
                j += 1
                final_quaters.append(i)    
    
    k = False
    print(final_quaters)
    print(quater_Q)
    
    # print(f"\nfilename {file} \n quaters{quaters}")
    # print(final_quaters)
    # print(df.columns)

    data_list = df.values.tolist()
    # print(f"Data list:- {data_list}")
    
    # print(f"\n After insert Data list:- {data_list}")
    param = []

    output_data = []
    dict = {}
            
    
    def isValid(data):
        return not ( data == '--')
    
    final_values = []
    i = 0
    for value in data_list:
        # if i == 0:
            parameter , *values = value
            # print("\n",values)
            for i in range(0,9):
                val = values[(i*5) : (i*5) + 5]
                val.reverse()
                # print("\n\n Parameter:-  ",parameter,"Val:- ",val)
                
                for i in val:
                    final_values.append(i)
        # else:
        #     break
    
    # print("\n\n Final value",len(final_values))
    # print(len(data_list))
    # print("\n\n Data List:- ",data_list[0])
    
    j = 0
    for i in range(0,len(data_list)):
        data = data_list[i]
        data_1  = data.copy()
        # print(data_1)
        
        data_1.clear()
        
        data_1.append(final_values[j:j+45])
        j += 45
       
        # print(data_1)
        flattend_list = data_1[0]
        # print(flattend_list)
        
        
        final_data = [data[0]] + flattend_list
        # print(final_data)
        data_list[i] = final_data
        # print("\n\n Data list:- ",data_list[i])
        
        
        # print("\n\n================================")
        # print(data_list)
        # final_data_list.append(data_list[i])
        # print("\n\n",final_data_list)
    

    for row in data_list:
        company, *values = row
        # print("\n Values inside row revsered:- ",value)
        if any(isValid(value) for value in values):
            output_data.append([company] + values)
            dict[company] = values

    # print("Output data before :-",output_data)
    # output_data = output_data[1:]
    # print("Output data after :-",output_data[1:])


    def isValids(val):
        if(str(val) == "nan" ):
            return True
        else:
            return False
        
    for row in output_data:
        paramter, *values = row
        # print("\nValue before:- ",values)
        #values.append("")
        if any(isValids(value) for value in values):
            param.append([""])
            param.append([paramter] + values)
        else:
            # print(values)
            float_list = [float(value.replace(',', '')) if value not in ["", "--"] else value for value in values]

            minus_float_list = [v for v in float_list if isinstance(v, float)]
            # print("\n\nMinus float 2:- ",minus_float_list,"\n")
            
            # max_val = max(value for value in minus_float_list)
            # min_val = min(value for value in minus_float_list)
            # avg = sum(minus_float_list) /len(minus_float_list)
            
            param.append([paramter] + values + [])

    
    param.insert(0,quater_Q)
    # print("\nData:- \n",param)

    col_name = col_names[0]

    print(col_name,quaters)
    output_file_path = f"{current_dir}/Companies/{sector_name}/{company_name}/Pruned_Excel/{file}"
    print("\n\nCurrent dir:- ",output_file_path)
    deleteExtraCell(output_file_path)
    output_df = pd.DataFrame(param,columns=[str(col_name)]+final_quaters)
    

    output_df.to_excel(output_file_path,index=False)
    print("\n\n\nSuccessfully saved the file in:- ",output_file_path)
    spacexcel(output_file_path,sector_name,company_name,file)


def deleteExtraCell(output_file_path):
    print("Inside delete cel; ",output_file_path)
    pass

def spacexcel(output_file_path,sector_name,company_name,file):
    global current_dir
    # file_path = r'C:\Users\sharm\OneDrive\Desktop\Kishan\Contractzy\WebScrapping\Tutorial\Code\7_Mar21_Mar22.xlsx' 
    file_path = f'{output_file_path}' 
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Increase column width and center-align numeric values
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Column letter for setting width

        for cell in col:
            # Center align numeric values
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center')
            
            # Calculate the maximum length of the cell values in the column
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set the column width slightly wider than the max length
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Save the modified Excel file
    output_path = f'{current_dir}/Companies/{sector_name}/{company_name}/Pruned_Excel/{file}'  # Choose your preferred output path
    print("Output path:- ",output_path)
    workbook.save(output_path)
    print("\n\n\nSuccessfully saved the file in:- ",output_path)


def prune_folders():
    global current_dir, path
    print("\n\n Path inside pruned folders:- ",path)
    sector_list = os.listdir(path)
    print("\nSector list:- ",sector_list)
    # print(len(sector_list))
    
    for i in range(len(sector_list)):
        path = f"{current_dir}/Companies"
        sector = sector_list[i]
        path = path + "/" + sector
        # print("Iteration:- ",i)
        # print("\nPath:- ",path)
        comapnies_list = os.listdir(path)
        # print("\nCompanies List:- ",comapnies_list)
        
        for i in range(len(comapnies_list)):
            company = comapnies_list[i]
            print(f"Company {i}:- {company}")
            # path = f"{current_dir}/Companies"
            path = f"{current_dir}/Companies/{sector}/{company}"
            print("\nFinal path to excel:-",path)

            prune_data(sector,company,"combined_excel_file.xlsx")
            # dir_list = os.listdir(path)
            
            # for dir in dir_list:
            #     print("\n\nFile name:- ",dir,"\n")
            #     prune_data(sector,company,dir)
    

if __name__ == '__main__':    
    
    prune_folders()
    
    # sector_name = "Refineries"
    # company_name  = "Bharat Petroleum Corporation Ltd"
    # file = "combined_excel_file.xlsx"
    
    # # path = f"{current_dir}/Companies/{sector_name}/{company_name}/Excel"
    # # kisu_list = os.listdir(path)
    # # #print(kisu_list)
    
    # # for kisu in kisu_list:
    # #     #print(kisu)
    # #     prune_data(sector_name,company_name,kisu)
    # prune_data(sector_name,company_name,file)
    
    
        
